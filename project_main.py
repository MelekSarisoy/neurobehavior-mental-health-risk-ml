# ─────────────────────────────────────────────
# 0. KÜTÜPHANELER
# ─────────────────────────────────────────────
import os, sys
import pandas as pd
import numpy as np
import warnings
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import seaborn as sns
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule

_THIS_DIR = os.path.dirname(os.path.abspath(__file__))
_DATA_CANDIDATES = [
    _THIS_DIR,
    os.path.join(_THIS_DIR, "data"),
    "/mnt/user-data/uploads",
    os.getcwd(),
]
_DATA_DIR = next(
    (d for d in _DATA_CANDIDATES if os.path.exists(os.path.join(d, "train.csv"))),
    _THIS_DIR
)
_OUT_DIR = _THIS_DIR

def dp(fname): return os.path.join(_DATA_DIR, fname)
def op(fname): return os.path.join(_OUT_DIR,  fname)

from sklearn.model_selection  import StratifiedKFold, cross_val_predict
from sklearn.preprocessing    import StandardScaler
from sklearn.metrics          import (
    accuracy_score, recall_score, precision_score,
    f1_score, matthews_corrcoef, confusion_matrix,
    roc_curve, auc
)
from sklearn.neighbors        import KNeighborsClassifier
from sklearn.naive_bayes      import GaussianNB
from sklearn.svm              import SVC
from sklearn.ensemble         import RandomForestClassifier
from sklearn.neural_network   import MLPClassifier
from sklearn.linear_model     import LogisticRegression
from sklearn.tree             import DecisionTreeClassifier

try:
    from xgboost import XGBClassifier
    XGB_OK = True
except ImportError:
    XGB_OK = False
    print("⚠  XGBoost bulunamadı → pip install xgboost")

warnings.filterwarnings("ignore")
BANNER = "═" * 68

# ─────────────────────────────────────────────
# 1. VERİ YÜKLEME
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 1 — Veri Yükleme")
print(BANNER)

train      = pd.read_csv(dp("train.csv"))
test       = pd.read_csv(dp("test.csv"))
main_df    = pd.read_csv(dp("NeuroBehavior-Clinical_Health_Risk_Sample_Dataset.csv"))
submission = pd.read_csv(dp("sample_submission.csv"))

TARGET = "Mental_Health_Risk"

print(f"  train.csv     : {train.shape[0]} satır × {train.shape[1]} sütun")
print(f"  test.csv      : {test.shape[0]} satır × {test.shape[1]} sütun")
print(f"  NeuroBehavior : {main_df.shape[0]} satır × {main_df.shape[1]} sütun")
print(f"\n  Hedef         : '{TARGET}'")
print(f"  Dağılım       : {train[TARGET].value_counts().to_dict()}")

# ─────────────────────────────────────────────
# 2. ÖN İŞLEME
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 2 — Ön İşleme")
print(BANNER)

def extract_hour(df):
    if "Timestamp" in df.columns:
        df = df.copy()
        df["Hour"] = pd.to_datetime(df["Timestamp"], errors="coerce").dt.hour.fillna(0).astype(int)
        df = df.drop(columns=["Timestamp"])
    return df

train = extract_hour(train)
test  = extract_hour(test)

train_ids = train["ID"].values
test_ids  = test["ID"].values

DROP_COLS    = ["ID"]
X_train_full = train.drop(columns=DROP_COLS + [TARGET])
y            = train[TARGET].values
X_test_full  = test.drop(columns=DROP_COLS)

X_train_full = X_train_full.fillna(X_train_full.median(numeric_only=True))
X_test_full  = X_test_full.fillna(X_train_full.median(numeric_only=True))

feature_names = X_train_full.columns.tolist()
print(f"  Özellikler ({len(feature_names)}): {feature_names}")

scaler     = StandardScaler()
X_scaled   = scaler.fit_transform(X_train_full)
X_test_sc  = scaler.transform(X_test_full)
X_raw      = X_train_full.values

SCALED_SET = {"LSVM", "RBF SVM", "MLP", "Logistic Reg."}

# ─────────────────────────────────────────────
# 3. MODELLER
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 3 — Modeller")
print(BANNER)

models = {
    "kNN"           : KNeighborsClassifier(n_neighbors=5),
    "Naive Bayes"   : GaussianNB(),
    "LSVM"          : SVC(kernel="linear", C=1.0, probability=True, random_state=42),
    "RBF SVM"       : SVC(kernel="rbf",    C=1.0, gamma="scale", probability=True, random_state=42),
    "Random Forest" : RandomForestClassifier(n_estimators=200, random_state=42),
    "MLP"           : MLPClassifier(hidden_layer_sizes=(128, 64), max_iter=600,
                                    early_stopping=True, random_state=42),
    "Decision Tree" : DecisionTreeClassifier(max_depth=8, random_state=42),
    "Logistic Reg." : LogisticRegression(max_iter=1000, random_state=42),
}
if XGB_OK:
    models["XGBoost"] = XGBClassifier(
        n_estimators=200, max_depth=6, learning_rate=0.05,
        eval_metric="logloss", random_state=42, verbosity=0
    )

for name in models:
    print(f"  ✔ {name}")

# ─────────────────────────────────────────────
# 4. YARDIMCI FONKSİYONLAR
# ─────────────────────────────────────────────

def specificity_score(y_true, y_pred):
    tn, fp, fn, tp = confusion_matrix(y_true, y_pred, labels=[0, 1]).ravel()
    return tn / (tn + fp) if (tn + fp) > 0 else 0.0

def evaluate(name, model, X_data, y_true, cv):
    y_pred = cross_val_predict(model, X_data, y_true, cv=cv, method="predict")
    acc  = accuracy_score (y_true, y_pred)
    rec  = recall_score   (y_true, y_pred, pos_label=1, zero_division=0)
    spec = specificity_score(y_true, y_pred)
    prec = precision_score(y_true, y_pred, pos_label=1, zero_division=0)
    f1   = f1_score       (y_true, y_pred, pos_label=1, zero_division=0)
    mcc  = matthews_corrcoef(y_true, y_pred)
    cm   = confusion_matrix(y_true, y_pred, labels=[0, 1])
    return {
        "Model"      : name,
        "Accuracy"   : round(acc  * 100, 2),
        "Recall"     : round(rec  * 100, 2),
        "Specificity": round(spec * 100, 2),
        "Precision"  : round(prec * 100, 2),
        "F1-Score"   : round(f1   * 100, 2),
        "MCC"        : round(mcc,         4),
        "_cm"        : cm,
        "_y_pred"    : y_pred,
    }

# ─────────────────────────────────────────────
# 5. 10-FOLD CV
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 4 — 10-Fold Stratified Cross Validation")
print(BANNER)

cv = StratifiedKFold(n_splits=10, shuffle=True, random_state=42)
results    = []
cms        = {}
pred_store = {}

for name, model in models.items():
    print(f"  ⏳ {name:<18} ...", end="\r")
    X_input = X_scaled if name in SCALED_SET else X_raw
    res = evaluate(name, model, X_input, y, cv)
    cms[name]        = res.pop("_cm")
    pred_store[name] = res.pop("_y_pred")
    results.append(res)
    print(f"  ✔ {name:<18}  Acc={res['Accuracy']:6.2f}%  F1={res['F1-Score']:6.2f}%  MCC={res['MCC']:+.4f}")

# ─────────────────────────────────────────────
# 6. SONUÇ TABLOSU (terminal)
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 5 — Sonuç Tablosu")
print(BANNER)

results_df = pd.DataFrame(results).set_index("Model")
header = f"{'Model':<18}" + "".join(f"  {c:>12}" for c in results_df.columns)
print(f"\n  {header}")
print("  " + "─" * len(header))
for mname, row in results_df.iterrows():
    vals = "".join(f"  {row[c]:>12.2f}" if c != "MCC" else f"  {row[c]:>+12.4f}"
                   for c in results_df.columns)
    print(f"  {mname:<18}{vals}")

print(f"\n  {'─'*50}")
for metric in ["Accuracy", "F1-Score", "MCC"]:
    best = results_df[metric].idxmax()
    val  = results_df.loc[best, metric]
    unit = "%" if metric != "MCC" else ""
    print(f"  🏆 En İyi {metric:<12}: {best:<18} → {val}{unit}")

# ─────────────────────────────────────────────
# 7. TEST TAHMİNİ + RİSK ETİKETİ
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 6 — Test Tahmini & Risk Etiketleme")
print(BANNER)

best_for_sub = results_df["F1-Score"].idxmax()
final_model  = models[best_for_sub]
X_sub_input  = X_scaled  if best_for_sub in SCALED_SET else X_raw
X_test_input = X_test_sc if best_for_sub in SCALED_SET else X_test_full.values

final_model.fit(X_sub_input, y)
test_preds = final_model.predict(X_test_input)

# Olasılık skoru (confidence)
try:
    test_proba = final_model.predict_proba(X_test_input)[:, 1]
except:
    test_proba = test_preds.astype(float)

# Risk seviyesi etiketi
def risk_label(pred, proba):
    if pred == 0:
        if proba < 0.25:
            return "🟢 Düşük Risk"
        else:
            return "🟡 Orta-Düşük Risk"
    else:
        if proba >= 0.75:
            return "🔴 Yüksek Risk"
        else:
            return "🟠 Orta-Yüksek Risk"

risk_labels  = [risk_label(p, pr) for p, pr in zip(test_preds, test_proba)]
risk_numeric = [
    "Düşük Risk" if p == 0 and pr < 0.25 else
    "Orta-Düşük Risk" if p == 0 else
    "Yüksek Risk" if pr >= 0.75 else
    "Orta-Yüksek Risk"
    for p, pr in zip(test_preds, test_proba)
]

test_result_df = test.copy()
test_result_df["Tahmin (0/1)"]      = test_preds
test_result_df["Risk Olasılığı"]    = np.round(test_proba, 4)
test_result_df["Risk Seviyesi"]     = risk_numeric

# Train için de tahmin
train_cv_preds = pred_store[best_for_sub]
train_proba_cv = cross_val_predict(
    models[best_for_sub], X_sub_input, y, cv=cv, method="predict_proba"
)[:, 1]
train_risk = [
    "Düşük Risk" if p == 0 and pr < 0.25 else
    "Orta-Düşük Risk" if p == 0 else
    "Yüksek Risk" if pr >= 0.75 else
    "Orta-Yüksek Risk"
    for p, pr in zip(train_cv_preds, train_proba_cv)
]
train_result_df = train.copy()
train_result_df["Tahmin (0/1)"]   = train_cv_preds
train_result_df["Risk Olasılığı"] = np.round(train_proba_cv, 4)
train_result_df["Risk Seviyesi"]  = train_risk
train_result_df["Gerçek Etiket"]  = ["Yüksek Risk" if v == 1 else "Düşük Risk" for v in y]
train_result_df["Doğru mu?"]      = train_cv_preds == y

print(f"  ✔ Kullanılan Model   : {best_for_sub}")
print(f"  Test Risk Dağılımı  :")
for lbl, cnt in pd.Series(risk_numeric).value_counts().items():
    print(f"    {lbl:<22}: {cnt} kişi")

# ─────────────────────────────────────────────
# 8. EXCEL RAPORU
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 7 — Excel Raporu Oluşturuluyor")
print(BANNER)

wb = Workbook()

# ── Renk Paleti ───────────────────────────────
C_HEADER_DARK  = "1F3864"   # koyu lacivert
C_HEADER_MID   = "2E75B6"   # orta mavi
C_HEADER_LIGHT = "BDD7EE"   # açık mavi
C_WHITE        = "FFFFFF"
C_GOLD         = "FFD700"
C_GREEN        = "E2EFDA"
C_RED_LIGHT    = "FCE4D6"
C_YELLOW_LIGHT = "FFEB9C"
C_ORANGE_LIGHT = "F4B942"
C_GRAY         = "F2F2F2"
C_BORDER       = "9DC3E6"

def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def header_cell(ws, row, col, value, bg=C_HEADER_DARK, fg=C_WHITE,
                bold=True, size=11, wrap=False, align="center"):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, color=fg, size=size)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center",
                            wrap_text=wrap)
    c.border    = thin_border()
    return c

def data_cell(ws, row, col, value, bg=C_WHITE, bold=False,
              align="center", number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Calibri", bold=bold, size=10)
    c.fill      = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal=align, vertical="center")
    c.border    = thin_border()
    if number_format:
        c.number_format = number_format
    return c

# ══════════════════════════════════════════════
# SAYFA 1 — Model Karşılaştırma Sonuçları
# ══════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Model Sonuçları"
ws1.sheet_view.showGridLines = False
ws1.column_dimensions["A"].width = 20

# Başlık bandı
ws1.merge_cells("A1:H1")
title_cell = ws1["A1"]
title_cell.value     = "NeuroBehavior Clinical Health Risk — Model Karşılaştırması (10-Fold CV)"
title_cell.font      = Font(name="Calibri", bold=True, size=15, color=C_WHITE)
title_cell.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:H2")
sub_cell = ws1["A2"]
sub_cell.value     = f"En İyi Model: {best_for_sub}  |  Değerlendirme: Stratified 10-Fold Cross Validation  |  Sınıf: Mental_Health_Risk (0/1)"
sub_cell.font      = Font(name="Calibri", size=10, color=C_WHITE, italic=True)
sub_cell.fill      = PatternFill("solid", fgColor=C_HEADER_MID)
sub_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[2].height = 20

# Sütun başlıkları
cols = ["Model", "Accuracy (%)", "Recall (%)", "Specificity (%)",
        "Precision (%)", "F1-Score (%)", "MCC", "Değerlendirme"]
col_widths = [20, 14, 12, 15, 14, 14, 10, 20]
for i, (col, w) in enumerate(zip(cols, col_widths), 1):
    header_cell(ws1, 3, i, col, bg=C_HEADER_MID, size=10, wrap=True)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[3].height = 30

# Her model için açıklama
def model_comment(acc):
    if acc >= 98: return "Mükemmel"
    if acc >= 95: return "Çok İyi"
    if acc >= 90: return "İyi"
    if acc >= 85: return "Orta"
    return "Geliştirilmeli"

best_acc_model = results_df["Accuracy"].idxmax()
for r_idx, (mname, row) in enumerate(results_df.iterrows(), 4):
    is_best = (mname == best_acc_model)
    bg = C_GOLD if is_best else (C_GRAY if r_idx % 2 == 0 else C_WHITE)
    bold = is_best

    data_cell(ws1, r_idx, 1, mname,            bg=bg, bold=bold, align="left")
    data_cell(ws1, r_idx, 2, row["Accuracy"],  bg=bg, bold=bold, number_format="0.00")
    data_cell(ws1, r_idx, 3, row["Recall"],    bg=bg, number_format="0.00")
    data_cell(ws1, r_idx, 4, row["Specificity"],bg=bg, number_format="0.00")
    data_cell(ws1, r_idx, 5, row["Precision"], bg=bg, number_format="0.00")
    data_cell(ws1, r_idx, 6, row["F1-Score"],  bg=bg, bold=bold, number_format="0.00")
    data_cell(ws1, r_idx, 7, row["MCC"],       bg=bg, number_format="0.0000")
    data_cell(ws1, r_idx, 8, model_comment(row["Accuracy"]), bg=bg, align="center")

ws1.row_dimensions[r_idx].height = 18

# Koşullu biçimlendirme — Accuracy sütunu
last_row = 3 + len(results_df)
ws1.conditional_formatting.add(
    f"B4:B{last_row}",
    ColorScaleRule(start_type="min", start_color="FCE4D6",
                   mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                   end_type="max", end_color="E2EFDA")
)
ws1.conditional_formatting.add(
    f"F4:F{last_row}",
    ColorScaleRule(start_type="min", start_color="FCE4D6",
                   mid_type="percentile", mid_value=50, mid_color="FFEB9C",
                   end_type="max", end_color="E2EFDA")
)

# En iyi satır notu
note_row = last_row + 2
ws1.merge_cells(f"A{note_row}:H{note_row}")
note = ws1.cell(row=note_row, column=1,
                value=f"★  En iyi model: {best_acc_model} "
                      f"(Accuracy: {results_df.loc[best_acc_model,'Accuracy']}%  |  "
                      f"F1: {results_df.loc[best_acc_model,'F1-Score']}%  |  "
                      f"MCC: {results_df.loc[best_acc_model,'MCC']})")
note.font      = Font(name="Calibri", bold=True, size=11, color="1F3864")
note.fill      = PatternFill("solid", fgColor="DEEAF1")
note.alignment = Alignment(horizontal="center", vertical="center")
note.border    = thin_border()
ws1.row_dimensions[note_row].height = 24

# Açıklama satırı (metrik tanımları)
desc_row = note_row + 1
ws1.merge_cells(f"A{desc_row}:H{desc_row}")
desc = ws1.cell(row=desc_row, column=1,
                value="Metrik Tanımları: Accuracy=Genel Doğruluk | Recall=Duyarlılık (TP/TP+FN) | "
                      "Specificity=Özgüllük (TN/TN+FP) | Precision=Hassasiyet | "
                      "F1=Harmonik Ortalama | MCC=Matthews Korelasyon Katsayısı")
desc.font      = Font(name="Calibri", size=9, italic=True, color="595959")
desc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[desc_row].height = 30

# SAYFA 2 — Test Seti Tahminleri + Risk
ws2 = wb.create_sheet("Test Tahminleri")
ws2.sheet_view.showGridLines = False

# Başlık
ws2.merge_cells("A1:P1")
t2 = ws2["A1"]
t2.value     = f"Test Seti Risk Tahminleri  —  Model: {best_for_sub}"
t2.font      = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
t2.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
t2.alignment = Alignment(horizontal="center", vertical="center")
ws2.row_dimensions[1].height = 28

# Sütun başlıkları
test_cols = list(test_result_df.columns)
col_w2 = {"ID":8, "Age":8, "Sleep_Hours":12, "Stress_Score":12,
           "Physical_Activity":17, "Screen_Time":12, "Workload_Level":14,
           "Mood_Stability":13, "Focus_Ability":12, "Social_Interaction":16,
           "Inflammation_Index":17, "Metabolic_Load":14, "Cortisol_Proxy":13,
           "Hour":8, "Tahmin (0/1)":13, "Risk Olasılığı":14, "Risk Seviyesi":18}

for ci, col in enumerate(test_cols, 1):
    header_cell(ws2, 2, ci, col, bg=C_HEADER_MID, size=9, wrap=True)
    ws2.column_dimensions[get_column_letter(ci)].width = col_w2.get(col, 12)
ws2.row_dimensions[2].height = 28

# Risk renkleri
RISK_COLORS = {
    "Düşük Risk"       : "E2EFDA",
    "Orta-Düşük Risk"  : "FFEB9C",
    "Orta-Yüksek Risk" : "FCE4D6",
    "Yüksek Risk"      : "FF7575",
}

for ri, (_, row_data) in enumerate(test_result_df.iterrows(), 3):
    risk_val = row_data["Risk Seviyesi"]
    row_bg   = RISK_COLORS.get(risk_val, C_WHITE)
    for ci, col in enumerate(test_cols, 1):
        val = row_data[col]
        fmt = "0.0000" if col == "Risk Olasılığı" else (
              "0.00"   if isinstance(val, float) and col not in ["Risk Seviyesi"] else None)
        align = "left" if col == "Risk Seviyesi" else "center"
        bg = row_bg if col in ["Tahmin (0/1)", "Risk Olasılığı", "Risk Seviyesi"] else (
             C_GRAY if ri % 2 == 0 else C_WHITE)
        data_cell(ws2, ri, ci, val, bg=bg, align=align, number_format=fmt)
    ws2.row_dimensions[ri].height = 16

# Risk istatistikleri bloğu (sağ tarafa)
stat_col = len(test_cols) + 2
header_cell(ws2, 2, stat_col,   "Risk Seviyesi",  bg=C_HEADER_DARK, size=10)
header_cell(ws2, 2, stat_col+1, "Kişi Sayısı",    bg=C_HEADER_DARK, size=10)
header_cell(ws2, 2, stat_col+2, "Oran (%)",        bg=C_HEADER_DARK, size=10)
ws2.column_dimensions[get_column_letter(stat_col)].width   = 20
ws2.column_dimensions[get_column_letter(stat_col+1)].width = 13
ws2.column_dimensions[get_column_letter(stat_col+2)].width = 12

risk_counts = pd.Series(risk_numeric).value_counts()
risk_order  = ["Düşük Risk","Orta-Düşük Risk","Orta-Yüksek Risk","Yüksek Risk"]
for si, rlbl in enumerate(risk_order, 3):
    cnt  = risk_counts.get(rlbl, 0)
    pct  = round(cnt / len(test_preds) * 100, 1)
    bg   = RISK_COLORS.get(rlbl, C_WHITE)
    data_cell(ws2, si, stat_col,   rlbl, bg=bg, align="left")
    data_cell(ws2, si, stat_col+1, cnt,  bg=bg)
    data_cell(ws2, si, stat_col+2, pct,  bg=bg, number_format="0.0")

# Toplam
tot_row = 3 + len(risk_order)
data_cell(ws2, tot_row, stat_col,   "TOPLAM",           bg=C_HEADER_LIGHT, bold=True)
data_cell(ws2, tot_row, stat_col+1, len(test_preds),     bg=C_HEADER_LIGHT, bold=True)
data_cell(ws2, tot_row, stat_col+2, 100.0,               bg=C_HEADER_LIGHT, bold=True, number_format="0.0")

# ══════════════════════════════════════════════
# SAYFA 3 — Train CV Sonuçları + Risk
# ══════════════════════════════════════════════
ws3 = wb.create_sheet("Train Tahminleri")
ws3.sheet_view.showGridLines = False

ws3.merge_cells("A1:R1")
t3 = ws3["A1"]
t3.value     = f"Train Seti CV Tahminleri & Risk Analizi  —  Model: {best_for_sub}"
t3.font      = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
t3.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
t3.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

tr_cols = list(train_result_df.columns)
for ci, col in enumerate(tr_cols, 1):
    header_cell(ws3, 2, ci, col, bg=C_HEADER_MID, size=9, wrap=True)
    ws3.column_dimensions[get_column_letter(ci)].width = col_w2.get(col, 14)
ws3.row_dimensions[2].height = 28

for ri, (_, row_data) in enumerate(train_result_df.iterrows(), 3):
    risk_val  = row_data["Risk Seviyesi"]
    is_wrong  = not row_data["Doğru mu?"]
    for ci, col in enumerate(tr_cols, 1):
        val   = row_data[col]
        fmt   = "0.0000" if col == "Risk Olasılığı" else (
                "0.00"   if isinstance(val, float) and col not in
                           ["Risk Seviyesi","Gerçek Etiket","Doğru mu?"] else None)
        align = "left"   if col in ["Risk Seviyesi","Gerçek Etiket"] else "center"
        if col == "Doğru mu?":
            bg = "E2EFDA" if val else "FCE4D6"
            val = "✔ Doğru" if val else "✘ Yanlış"
        elif col == "Risk Seviyesi":
            bg = RISK_COLORS.get(risk_val, C_WHITE)
        elif col in ["Tahmin (0/1)","Risk Olasılığı","Gerçek Etiket"]:
            bg = RISK_COLORS.get(risk_val, C_WHITE)
        else:
            bg = C_GRAY if ri % 2 == 0 else C_WHITE
        data_cell(ws3, ri, ci, val, bg=bg, align=align, number_format=fmt)
    ws3.row_dimensions[ri].height = 16

# ══════════════════════════════════════════════
# SAYFA 4 — Özet & İstatistikler
# ══════════════════════════════════════════════
ws4 = wb.create_sheet("Özet & İstatistikler")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 28
ws4.column_dimensions["B"].width = 22
ws4.column_dimensions["C"].width = 18
ws4.column_dimensions["D"].width = 18

ws4.merge_cells("A1:D1")
t4 = ws4["A1"]
t4.value     = "Proje Özeti — NeuroBehavior Mental Health Risk Sınıflandırması"
t4.font      = Font(name="Calibri", bold=True, size=14, color=C_WHITE)
t4.fill      = PatternFill("solid", fgColor=C_HEADER_DARK)
t4.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 30

# Genel bilgiler
info_rows = [
    ("Veri Seti",             "NeuroBehavior Clinical Health Risk"),
    ("Eğitim Örnekleri",      f"{len(train)} adet"),
    ("Test Örnekleri",        f"{len(test)} adet"),
    ("Özellik Sayısı",        f"{len(feature_names)} adet"),
    ("Hedef Değişken",        "Mental_Health_Risk (Binary: 0/1)"),
    ("Doğrulama Yöntemi",     "Stratified 10-Fold Cross Validation"),
    ("Toplam Model Sayısı",   f"{len(models)} model"),
    ("En İyi Model",          best_for_sub),
    ("En İyi Accuracy",       f"%{results_df['Accuracy'].max():.2f}"),
    ("En İyi F1-Score",       f"%{results_df['F1-Score'].max():.2f}"),
    ("En İyi MCC",            f"{results_df['MCC'].max():.4f}"),
    ("AUC-ROC (En İyi)",      "Ayrıca hesaplanmıştır"),
]

ws4.merge_cells("A3:D3")
sec1 = ws4["A3"]
sec1.value     = "📊  Genel Proje Bilgileri"
sec1.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
sec1.fill      = PatternFill("solid", fgColor=C_HEADER_MID)
sec1.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[3].height = 22

for ri, (k, v) in enumerate(info_rows, 4):
    bg = C_GRAY if ri % 2 == 0 else C_WHITE
    data_cell(ws4, ri, 1, k, bg=bg, bold=True,  align="left")
    data_cell(ws4, ri, 2, v, bg=bg, bold=False, align="left")
    ws4.row_dimensions[ri].height = 18

# Risk dağılımı özeti (test)
sec2_row = 4 + len(info_rows) + 1
ws4.merge_cells(f"A{sec2_row}:D{sec2_row}")
sec2 = ws4[f"A{sec2_row}"]
sec2.value     = "🎯  Test Seti Risk Dağılımı"
sec2.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
sec2.fill      = PatternFill("solid", fgColor=C_HEADER_MID)
sec2.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[sec2_row].height = 22

header_cell(ws4, sec2_row+1, 1, "Risk Seviyesi",  bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, sec2_row+1, 2, "Kişi Sayısı",    bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, sec2_row+1, 3, "Oran (%)",        bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, sec2_row+1, 4, "Açıklama",        bg=C_HEADER_LIGHT, fg="1F3864")

RISK_DESC = {
    "Düşük Risk"       : "Model güvenle düşük risk tahmin etti",
    "Orta-Düşük Risk"  : "Düşük risk, ancak izleme önerilir",
    "Orta-Yüksek Risk" : "Yüksek risk eğilimi, dikkat gerekli",
    "Yüksek Risk"      : "Model güvenle yüksek risk tahmin etti",
}

for si, rlbl in enumerate(risk_order, sec2_row+2):
    cnt = risk_counts.get(rlbl, 0)
    pct = round(cnt / len(test_preds) * 100, 1)
    bg  = RISK_COLORS.get(rlbl, C_WHITE)
    data_cell(ws4, si, 1, rlbl,              bg=bg, align="left",   bold=True)
    data_cell(ws4, si, 2, cnt,               bg=bg, align="center")
    data_cell(ws4, si, 3, pct,               bg=bg, align="center", number_format="0.0")
    data_cell(ws4, si, 4, RISK_DESC[rlbl],   bg=bg, align="left")
    ws4.row_dimensions[si].height = 18

# Model sıralaması (Accuracy'e göre)
rank_start = sec2_row + len(risk_order) + 4
ws4.merge_cells(f"A{rank_start}:D{rank_start}")
sec3 = ws4[f"A{rank_start}"]
sec3.value     = "🏆  Model Sıralaması (Accuracy)"
sec3.font      = Font(name="Calibri", bold=True, size=11, color=C_WHITE)
sec3.fill      = PatternFill("solid", fgColor=C_HEADER_MID)
sec3.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[rank_start].height = 22

header_cell(ws4, rank_start+1, 1, "Sıra",       bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, rank_start+1, 2, "Model",       bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, rank_start+1, 3, "Accuracy (%)",bg=C_HEADER_LIGHT, fg="1F3864")
header_cell(ws4, rank_start+1, 4, "F1-Score (%)",bg=C_HEADER_LIGHT, fg="1F3864")

ranked = results_df.sort_values("Accuracy", ascending=False)
medal  = {0:"🥇", 1:"🥈", 2:"🥉"}
for ri, (mname, row) in enumerate(ranked.iterrows()):
    rr  = rank_start + 2 + ri
    bg  = C_GOLD if ri == 0 else (C_GRAY if ri % 2 == 0 else C_WHITE)
    data_cell(ws4, rr, 1, f"{medal.get(ri, str(ri+1)+'.')} {ri+1}", bg=bg, bold=(ri==0))
    data_cell(ws4, rr, 2, mname,           bg=bg, bold=(ri==0), align="left")
    data_cell(ws4, rr, 3, row["Accuracy"], bg=bg, bold=(ri==0), number_format="0.00")
    data_cell(ws4, rr, 4, row["F1-Score"], bg=bg, bold=(ri==0), number_format="0.00")
    ws4.row_dimensions[rr].height = 18

# Excel kaydet
excel_path = op("ML_Sonuclar_Raporu.xlsx")
wb.save(excel_path)
print(f"  ✔ Excel raporu kaydedildi → ML_Sonuclar_Raporu.xlsx")

# ─────────────────────────────────────────────
# 9. GÖRSELLEŞTİRMELER
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  ADIM 8 — Görseller")
print(BANNER)

metric_cols = ["Accuracy","Recall","Specificity","Precision","F1-Score","MCC"]
colors = sns.color_palette("tab10", len(results_df))

fig, axes = plt.subplots(2, 3, figsize=(20, 11))
fig.suptitle("NeuroBehavior — Model Karşılaştırması (10-Fold CV)",
             fontsize=14, fontweight="bold")
for ax, metric in zip(axes.flatten(), metric_cols):
    vals = results_df[metric]
    bars = ax.barh(results_df.index, vals, color=colors, edgecolor="white", height=0.6)
    ax.set_title(metric, fontweight="bold", fontsize=12)
    is_mcc = (metric == "MCC")
    ax.set_xlabel("MCC Skoru" if is_mcc else "%", fontsize=10)
    ax.set_xlim(0 if not is_mcc else min(-0.1, vals.min()*1.1),
                100 if not is_mcc else max(1.05, vals.max()*1.1))
    ax.axvline(0, color="gray", linewidth=0.5)
    best_idx = list(results_df.index).index(vals.idxmax())
    for i, (bar, val) in enumerate(zip(bars, vals)):
        ax.text(bar.get_width() + (0.5 if not is_mcc else 0.01),
                bar.get_y() + bar.get_height()/2,
                f"{val:.2f}", va="center", fontsize=8.5)
        if i == best_idx:
            bar.set_edgecolor("gold"); bar.set_linewidth(2.5)
plt.tight_layout()
plt.savefig(op("01_model_comparison_bars.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 01_model_comparison_bars.png")

# Isıl Harita
fig, ax = plt.subplots(figsize=(11, 6))
hm_norm = results_df.copy()
hm_norm["MCC"] = (hm_norm["MCC"] + 1) / 2 * 100
sns.heatmap(hm_norm, annot=results_df.values, fmt=".2f", cmap="YlOrRd",
            linewidths=0.6, linecolor="white", ax=ax,
            cbar_kws={"label": "Normalize Skor (%)"})
ax.set_title("Performans Isıl Haritası — 10-Fold CV", fontsize=12, fontweight="bold")
plt.tight_layout()
plt.savefig(op("02_heatmap_results.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 02_heatmap_results.png")

# Karışıklık Matrisleri
n_models = len(models)
cols_cm  = min(3, n_models)
rows_cm  = (n_models + cols_cm - 1) // cols_cm
fig, axes = plt.subplots(rows_cm, cols_cm, figsize=(5.5*cols_cm, 4.5*rows_cm))
axes_flat = axes.flatten() if n_models > 1 else [axes]
fig.suptitle("Karışıklık Matrisleri (10-Fold CV)", fontsize=14, fontweight="bold")
for ax, (name, cm) in zip(axes_flat, cms.items()):
    sns.heatmap(cm, annot=True, fmt="d", cmap="Blues",
                xticklabels=["0\n(Düşük)","1\n(Yüksek)"],
                yticklabels=["0\n(Düşük)","1\n(Yüksek)"],
                ax=ax, cbar=False)
    ax.set_title(f"{name}\n(Acc: {results_df.loc[name,'Accuracy']:.2f}%)", fontweight="bold")
    ax.set_xlabel("Tahmin"); ax.set_ylabel("Gerçek")
for ax in axes_flat[len(cms):]:
    ax.axis("off")
plt.tight_layout()
plt.savefig(op("03_confusion_matrices.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 03_confusion_matrices.png")

# Radar
radar_metrics = ["Accuracy","Recall","Specificity","Precision","F1-Score"]
radar_data    = results_df[radar_metrics].copy()
angles = np.linspace(0, 2*np.pi, len(radar_metrics), endpoint=False).tolist()
angles += angles[:1]
fig, ax = plt.subplots(figsize=(9,9), subplot_kw={"polar":True})
ax.set_theta_offset(np.pi/2); ax.set_theta_direction(-1)
ax.set_thetagrids(np.degrees(angles[:-1]), radar_metrics, fontsize=11)
cmap_r = plt.cm.tab10; patches = []
for i, (mname, row) in enumerate(radar_data.iterrows()):
    vals_r = row.tolist() + [row.tolist()[0]]
    c = cmap_r(i/len(radar_data))
    ax.plot(angles, vals_r, linewidth=1.8, color=c)
    ax.fill(angles, vals_r, alpha=0.07, color=c)
    patches.append(mpatches.Patch(color=c, label=mname))
ax.set_ylim(0, 100)
ax.set_title("Model Radar Karşılaştırması\n", fontsize=13, fontweight="bold")
ax.legend(handles=patches, loc="lower right", bbox_to_anchor=(1.35,-0.05), fontsize=9)
plt.tight_layout()
plt.savefig(op("04_radar_chart.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 04_radar_chart.png")

# ROC
best_by_f1     = results_df["F1-Score"].idxmax()
best_model_obj = models[best_by_f1]
X_best = X_scaled if best_by_f1 in SCALED_SET else X_raw
try:
    y_prob_cv = cross_val_predict(best_model_obj, X_best, y, cv=cv, method="predict_proba")[:,1]
    fpr, tpr, _ = roc_curve(y, y_prob_cv)
    roc_auc = auc(fpr, tpr)
    fig, ax = plt.subplots(figsize=(7,6))
    ax.plot(fpr, tpr, lw=2.5, color="steelblue", label=f"{best_by_f1}\nAUC = {roc_auc:.4f}")
    ax.plot([0,1],[0,1],"k--",lw=1,label="Rastgele")
    ax.set_xlim([0,1]); ax.set_ylim([0,1.02])
    ax.set_xlabel("FPR",fontsize=12); ax.set_ylabel("TPR",fontsize=12)
    ax.set_title(f"ROC Eğrisi — {best_by_f1}", fontsize=13, fontweight="bold")
    ax.legend(fontsize=11); ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(op("05_roc_curve_best.png"), dpi=150, bbox_inches="tight")
    plt.close()
    print(f"  ✔ 05_roc_curve_best.png  (AUC = {roc_auc:.4f})")
except Exception as e:
    print(f"  ⚠  ROC grafiği oluşturulamadı: {e}")

# Özellik Önemi
rf_model = RandomForestClassifier(n_estimators=200, random_state=42)
rf_model.fit(X_raw, y)
importances = pd.Series(rf_model.feature_importances_, index=feature_names).sort_values(ascending=True)
fig, ax = plt.subplots(figsize=(9,6))
colors_fi = ["#e74c3c" if v >= importances.quantile(0.75) else "#3498db" for v in importances.values]
importances.plot(kind="barh", ax=ax, color=colors_fi, edgecolor="white")
ax.set_title("Özellik Önemi — Random Forest\n(Kırmızı: En Önemli %25)", fontsize=13, fontweight="bold")
ax.set_xlabel("Gini Önem Skoru")
ax.axvline(importances.quantile(0.75), color="crimson", linestyle="--", lw=1.5, label="75. Yüzdelik")
ax.legend(fontsize=10)
plt.tight_layout()
plt.savefig(op("06_feature_importance.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 06_feature_importance.png")

# Risk Pasta Grafiği
fig, axes = plt.subplots(1, 2, figsize=(13, 6))
fig.suptitle("Risk Dağılımı Analizi", fontsize=13, fontweight="bold")
risk_color_list = ["#2ecc71","#f1c40f","#e67e22","#e74c3c"]

# Test pasta
rc_test = [risk_counts.get(r, 0) for r in risk_order]
axes[0].pie(rc_test, labels=risk_order, autopct="%1.1f%%",
            colors=risk_color_list, startangle=140,
            wedgeprops={"edgecolor":"white","linewidth":1.5})
axes[0].set_title(f"Test Seti Risk Dağılımı\n(n={len(test_preds)})", fontweight="bold")

# Train pasta
train_risk_counts = pd.Series(train_risk).value_counts()
rc_train = [train_risk_counts.get(r, 0) for r in risk_order]
axes[1].pie(rc_train, labels=risk_order, autopct="%1.1f%%",
            colors=risk_color_list, startangle=140,
            wedgeprops={"edgecolor":"white","linewidth":1.5})
axes[1].set_title(f"Train Seti Risk Dağılımı\n(10-Fold CV, n={len(y)})", fontweight="bold")

plt.tight_layout()
plt.savefig(op("07_risk_distribution.png"), dpi=150, bbox_inches="tight")
plt.close()
print("  ✔ 07_risk_distribution.png")

# Submission CSV
submission_out = pd.DataFrame({"ID": test_ids, "Mental_Health_Risk": test_preds})
submission_out.to_csv(op("submission_predictions.csv"), index=False)

# ─────────────────────────────────────────────
# 10. ÖZET
# ─────────────────────────────────────────────
print(f"\n{BANNER}")
print("  PROJE TAMAMLANDI")
print(BANNER)
print(f"""

  Risk Seviyeleri:
    🟢 Düşük Risk       → Tahmin=0, Olasılık < %25
    🟡 Orta-Düşük Risk  → Tahmin=0, Olasılık ≥ %25
    🟠 Orta-Yüksek Risk → Tahmin=1, Olasılık < %75
    🔴 Yüksek Risk      → Tahmin=1, Olasılık ≥ %75
""")